"""
Excel Integration Module
Handles reading from and writing to Excel files
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
from pathlib import Path
from typing import List, Dict, Optional, Tuple
import numpy as np


class ExcelHandler:
    """Handles Excel file operations for ledger classification"""
    
    def __init__(self, config_loader):
        """
        Initialize Excel handler
        
        Args:
            config_loader: ConfigLoader instance
        """
        self.config = config_loader
        self.excel_config = config_loader.get_excel_config()
    
    def read_excel(self, file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Read Excel file
        
        Args:
            file_path: Path to Excel file
            sheet_name: Optional sheet name (uses first sheet if None)
        
        Returns:
            DataFrame with Excel data
        """
        if not Path(file_path).exists():
            raise FileNotFoundError(f"Excel file not found: {file_path}")
        
        try:
            if sheet_name:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                df = pd.read_excel(file_path)
            
            return df
        except Exception as e:
            raise ValueError(f"Error reading Excel file: {e}")
    
    def validate_columns(self, df: pd.DataFrame, required_columns: List[str]) -> bool:
        """
        Validate that required columns exist
        
        Args:
            df: Input DataFrame
            required_columns: List of required column names
        
        Returns:
            True if all columns exist
        """
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")
        
        return True
    
    def prepare_data_for_classification(self,
                                       df: pd.DataFrame,
                                       classification_level: int) -> Tuple[pd.DataFrame, List[int]]:
        """
        Prepare data for classification
        
        Args:
            df: Input DataFrame
            classification_level: Target classification level (3 or 4)
        
        Returns:
            Tuple of (prepared DataFrame, indices to classify)
        """
        input_col = self.excel_config['input_column']
        
        # Validate input column exists
        if input_col not in df.columns:
            raise ValueError(f"Input column '{input_col}' not found in Excel file")
        
        # For level 4, also need classification 3
        if classification_level == 4:
            if 'Classification 3' not in df.columns:
                raise ValueError("Classification 3 column required for level 4 classification")
        
        # Find rows that need classification (non-empty ledger name, empty classification)
        output_col = f'Classification {classification_level}'
        
        # Add output column if it doesn't exist
        if output_col not in df.columns:
            df[output_col] = ''
        
        # Add confidence column
        confidence_col = f'{output_col} Confidence'
        if confidence_col not in df.columns:
            df[confidence_col] = np.nan
        
        # Find indices to classify
        indices_to_classify = df[
            (df[input_col].notna()) & 
            (df[input_col].astype(str).str.strip() != '') &
            ((df[output_col].isna()) | (df[output_col].astype(str).str.strip() == ''))
        ].index.tolist()
        
        return df, indices_to_classify
    
    def write_classifications(self,
                            df: pd.DataFrame,
                            file_path: str,
                            classification_level: int,
                            highlight_low_confidence: bool = True) -> None:
        """
        Write classifications back to Excel with formatting
        
        Args:
            df: DataFrame with classifications
            file_path: Output file path
            classification_level: Classification level (3 or 4)
            highlight_low_confidence: Whether to highlight low confidence predictions
        """
        output_col = f'Classification {classification_level}'
        confidence_col = f'{output_col} Confidence'
        confidence_threshold = self.excel_config.get('confidence_threshold', 0.7)
        
        # Write to Excel
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Classified Ledgers', index=False)
            
            if highlight_low_confidence:
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Classified Ledgers']
                
                # Define fill colors
                low_confidence_fill = PatternFill(
                    start_color='FFC7CE',
                    end_color='FFC7CE',
                    fill_type='solid'
                )
                high_confidence_fill = PatternFill(
                    start_color='C6EFCE',
                    end_color='C6EFCE',
                    fill_type='solid'
                )
                
                # Find column indices
                output_col_idx = df.columns.get_loc(output_col) + 1
                confidence_col_idx = df.columns.get_loc(confidence_col) + 1
                
                # Apply formatting
                for row_idx, confidence in enumerate(df[confidence_col], start=2):
                    if pd.notna(confidence):
                        if confidence < confidence_threshold:
                            # Low confidence - red
                            worksheet.cell(row_idx, output_col_idx).fill = low_confidence_fill
                            worksheet.cell(row_idx, confidence_col_idx).fill = low_confidence_fill
                        else:
                            # High confidence - green
                            worksheet.cell(row_idx, output_col_idx).fill = high_confidence_fill
                            worksheet.cell(row_idx, confidence_col_idx).fill = high_confidence_fill
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        print(f"Classifications written to: {file_path}")
    
    def create_classification_report(self,
                                    df: pd.DataFrame,
                                    file_path: str,
                                    classification_level: int) -> None:
        """
        Create a summary report of classifications
        
        Args:
            df: DataFrame with classifications
            file_path: Output file path
            classification_level: Classification level (3 or 4)
        """
        output_col = f'Classification {classification_level}'
        confidence_col = f'{output_col} Confidence'
        
        # Create summary statistics
        summary = {
            'Total Ledgers': len(df),
            'Classified': df[output_col].notna().sum(),
            'Unclassified': df[output_col].isna().sum(),
            'Average Confidence': df[confidence_col].mean(),
            'Min Confidence': df[confidence_col].min(),
            'Max Confidence': df[confidence_col].max()
        }
        
        # Class distribution
        class_distribution = df[output_col].value_counts().to_dict()
        
        # Confidence statistics by class
        confidence_by_class = df.groupby(output_col)[confidence_col].agg([
            'mean', 'min', 'max', 'count'
        ]).to_dict('index')
        
        # Write report to Excel
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            # Summary sheet
            summary_df = pd.DataFrame([summary])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Class distribution sheet
            dist_df = pd.DataFrame(
                list(class_distribution.items()),
                columns=['Classification', 'Count']
            )
            dist_df['Percentage'] = (dist_df['Count'] / dist_df['Count'].sum() * 100).round(2)
            dist_df.to_excel(writer, sheet_name='Class Distribution', index=False)
            
            # Confidence by class sheet
            conf_df = pd.DataFrame.from_dict(confidence_by_class, orient='index')
            conf_df.reset_index(inplace=True)
            conf_df.columns = ['Classification', 'Avg Confidence', 'Min Confidence', 'Max Confidence', 'Count']
            conf_df.to_excel(writer, sheet_name='Confidence by Class', index=False)
            
            # Low confidence predictions
            threshold = self.excel_config.get('confidence_threshold', 0.7)
            low_conf_df = df[df[confidence_col] < threshold][[
                self.excel_config['input_column'],
                output_col,
                confidence_col
            ]].copy()
            low_conf_df.to_excel(writer, sheet_name='Low Confidence', index=False)
        
        print(f"Classification report saved to: {file_path}")
    
    def export_training_data(self,
                            df: pd.DataFrame,
                            file_path: str,
                            classification_level: int) -> None:
        """
        Export data in format suitable for training
        
        Args:
            df: DataFrame with classifications
            file_path: Output file path
            classification_level: Classification level (3 or 4)
        """
        input_col = self.excel_config['input_column']
        output_col = f'Classification {classification_level}'
        
        # Select required columns
        if classification_level == 4:
            export_df = df[[input_col, 'Classification 3', output_col]].copy()
        else:
            export_df = df[[input_col, output_col]].copy()
        
        # Remove rows with missing values
        export_df = export_df.dropna()
        
        # Save to Excel
        export_df.to_excel(file_path, index=False)
        print(f"Training data exported to: {file_path}")


if __name__ == "__main__":
    from config_loader import ConfigLoader
    
    config = ConfigLoader()
    handler = ExcelHandler(config)
    print("Excel handler initialized successfully!")
